import csv as _csv
import io
import json
import os
import re
import socket
import urllib.parse
import urllib.request
import webbrowser
from datetime import datetime
from io import BytesIO

from flask import Flask, Response, jsonify, render_template, request

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO = os.path.join(BASE_DIR, "tabeladecompras.xlsx")
HISTORICO = os.path.join(BASE_DIR, "historico_precos.json")
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
HEADERS = ["Peça", "Nome", "Valor", "URL", "Notas", "Imagem"]

UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")


def carregar_env():
    env_path = os.path.join(BASE_DIR, ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as fh:
        for linha in fh:
            linha = linha.strip()
            if not linha or linha.startswith("#") or "=" not in linha:
                continue
            chave, _, valor = linha.partition("=")
            os.environ.setdefault(chave.strip(), valor.strip().strip('"').strip("'"))


carregar_env()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024


def parse_valor(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    text = re.sub(r"[^\d.]", "", text)
    return float(text) if text else 0.0


def format_valor(value):
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def load_items():
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb["Sheet1"]
    items = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, 7)]
        if all(v is None or str(v).strip() == "" for v in values[:4]):
            continue
        item = {
            "id": row - 2,
            "peca": values[0] or "",
            "nome": values[1] or "",
            "valor": parse_valor(values[2]),
            "valor_formatado": "",
            "url": values[3] or "",
            "nota": values[4] or "",
            "imagem": values[5] or "",
        }
        item["valor_formatado"] = format_valor(item["valor"])
        items.append(item)
    return items


def save_items(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for item in items:
        ws.append([item["peca"], item["nome"], item["valor"], item["url"], item.get("nota", ""), item.get("imagem", "")])
        ws.cell(row=ws.max_row, column=3).number_format = 'R$ #,##0.00'
    wb.save(ARQUIVO)


def validate(payload):
    peca = (payload.get("peca") or "").strip()
    nome = (payload.get("nome") or "").strip()
    try:
        valor = float(str(payload.get("valor", "")).strip().replace("R$", "").replace(" ", "").replace(".", "").replace(",", "."))
    except (TypeError, ValueError):
        valor = 0.0
    url = (payload.get("url") or "").strip()
    nota = (payload.get("nota") or "").strip()
    imagem = (payload.get("imagem") or "").strip()
    if not nome:
        raise ValueError("O campo Nome é obrigatório.")
    if valor <= 0:
        raise ValueError("O valor deve ser maior que zero.")
    return peca, nome, valor, url, nota, imagem


def load_historico():
    if not os.path.exists(HISTORICO):
        return []
    try:
        with open(HISTORICO, encoding="utf-8") as fh:
            dados = json.load(fh)
        return dados if isinstance(dados, list) else []
    except (OSError, ValueError):
        return []


def salvar_historico(historico):
    with open(HISTORICO, "w", encoding="utf-8") as fh:
        json.dump(historico, fh, ensure_ascii=False, indent=2)


def registrar_historico(nome, peca, valor_antigo, valor_novo):
    if round(valor_antigo, 2) == round(valor_novo, 2):
        return
    historico = load_historico()
    historico.insert(0, {
        "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "peca": peca,
        "nome": nome,
        "valor_antigo": valor_antigo,
        "valor_novo": valor_novo,
    })
    salvar_historico(historico[:100])


def load_config():
    padrao = {"cep": "", "orcamento": 0.0, "ml_token": ""}
    try:
        with open(CONFIG_FILE, encoding="utf-8") as fh:
            cfg = json.load(fh)
        if isinstance(cfg, dict):
            for chave in padrao:
                if chave in cfg:
                    padrao[chave] = cfg[chave]
    except (OSError, ValueError):
        pass
    padrao["cep"] = padrao.get("cep") or ""
    try:
        padrao["orcamento"] = float(padrao.get("orcamento") or 0)
    except (TypeError, ValueError):
        padrao["orcamento"] = 0.0
    return padrao


def salvar_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as fh:
        json.dump(cfg, fh, ensure_ascii=False, indent=2)


def http_get(url, timeout=8, token=None):
    headers = {
        "User-Agent": UA,
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/json,text/plain,*/*;q=0.8",
    }
    if token:
        headers["Authorization"] = "Bearer " + token
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8", "ignore")


def http_post_form(url, campos, timeout=10):
    data = urllib.parse.urlencode(campos).encode()
    req = urllib.request.Request(
        url,
        data=data,
        headers={"User-Agent": UA, "Content-Type": "application/x-www-form-urlencoded"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode("utf-8", "ignore"))


def ml_token_atual():
    return os.environ.get("ML_ACCESS_TOKEN") or load_config().get("ml_token", "")


def frete_mercadolivre(link, cep):
    if not cep or not link:
        return None
    m = re.search(r"(MLB\d+)", link)
    if not m:
        return None
    try:
        dados = json.loads(http_get(
            f"https://api.mercadolibre.com/items/{m.group(1)}/shipping_options?zip_code={cep}",
            timeout=5,
        ))
        opcoes = dados.get("options") or []
        if not opcoes:
            return None
        custo = min(o.get("cost", None) for o in opcoes if o.get("cost") is not None)
        return custo
    except Exception:
        return None


def buscar_mercadolivre_api(q, cep):
    token = ml_token_atual()
    url = ("https://api.mercadolibre.com/sites/MLB/search"
           f"?q={urllib.parse.quote(q)}&limit=8")
    dados = json.loads(http_get(url, token=token))
    resultados = []
    for it in dados.get("results", []):
        preco = it.get("price")
        if preco is None:
            continue
        permalink = (it.get("permalink") or "").split("?")[0]
        frete_valor = frete_mercadolivre(permalink, cep)
        thumb = it.get("thumbnail")
        resultados.append({
            "loja": "Mercado Livre",
            "nome": it.get("title") or "",
            "preco": float(preco),
            "valor_formatado": format_valor(float(preco)),
            "link": permalink,
            "imagem": thumb if thumb and thumb.startswith("http") else None,
            "frete": format_valor(frete_valor) if frete_valor is not None else None,
            "frete_valor": frete_valor,
        })
    return resultados


def buscar_shopee_api(q, cep):
    url = ("https://shopee.com.br/api/v4/search/search_items"
           f"?by=relevancy&keyword={urllib.parse.quote(q)}&limit=12&newest=0"
           "&order=desc&page_type=search&scenario=PAGE_GLOBAL_SEARCH&version=2")
    dados = json.loads(http_get(url))
    resultados = []
    for it in (dados.get("items") or [])[:12]:
        ib = it.get("item_basic") or {}
        nome = ib.get("name")
        preco_cent = ib.get("price")
        if not nome or preco_cent is None:
            continue
        preco = preco_cent / 100.0
        itemid, shopid = ib.get("itemid"), ib.get("shopid")
        link = f"https://shopee.com.br/product/{shopid}/{itemid}" if shopid and itemid else None
        img = ib.get("image")
        resultados.append({
            "loja": "Shopee",
            "nome": nome,
            "preco": preco,
            "valor_formatado": format_valor(preco),
            "link": link or "https://shopee.com.br/search?keyword=" + urllib.parse.quote(q),
            "imagem": ("https://down-br.img.susercontent.com.br/" + img) if img else None,
            "frete": None,
            "frete_valor": None,
        })
    return resultados


def loja_por_dominio(host):
    host = (host or "").lower().replace("www.", "")
    if "mercadolivre" in host or "meli" in host:
        return "Mercado Livre"
    if "shopee" in host:
        return "Shopee"
    if "amazon" in host:
        return "Amazon"
    if "aliexpress" in host:
        return "AliExpress"
    if "kabum" in host:
        return "KaBum!"
    if "terabyteshop" in host or "terabyte" in host:
        return "Terabyte"
    if "pichau" in host:
        return "Pichau"
    if "firge" in host:
        return "Firge"
    return host or "Loja"


def buscar_ddg(q, cep=""):
    url = "https://lite.duckduckgo.com/lite/?q=" + urllib.parse.quote(q)
    html = http_get(url, timeout=10)
    resultados = []
    vistos = set()
    for m in re.finditer(r'href="(?:/l/|//duckduckgo\.com/l/)\?uddg=([^"&]+)[^"]*"[^>]*>(.*?)</a>', html, re.S):
        href = urllib.parse.unquote(m.group(1))
        titulo = re.sub(r"<[^>]+>", "", m.group(2)).strip()
        if not href.startswith("http") or "duckduckgo.com/y.js" in href:
            continue
        if href in vistos:
            continue
        vistos.add(href)
        resultados.append({
            "loja": loja_por_dominio(urllib.parse.urlparse(href).netloc),
            "nome": titulo or href,
            "preco": None,
            "valor_formatado": None,
            "link": href,
            "imagem": None,
            "frete": None,
            "frete_valor": None,
        })
        if len(resultados) >= 8:
            break
    return resultados


def arquivo_para_items(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    items = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, 7)]
        if all(v is None or str(v).strip() == "" for v in values[:4]):
            continue
        nome = str(values[1] or "").strip()
        if not nome:
            raise ValueError("Há uma linha com peça sem nome (linha %d)." % row)
        items.append({
            "peca": str(values[0] or "").strip(),
            "nome": nome,
            "valor": parse_valor(values[2]),
            "url": str(values[3] or "").strip(),
            "nota": str(values[4] or "").strip(),
            "imagem": str(values[5] or "").strip(),
        })
    if not items:
        raise ValueError("A planilha importada está vazia.")
    return items


@app.after_request
def no_cache(response):
    if request.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
    return response


@app.route("/")
def index():
    items = load_items()
    total = sum(i["valor"] for i in items)
    return render_template("index.html", items=items, total=format_valor(total))


@app.route("/favicon.ico")
def favicon():
    return "", 204


@app.route("/api/items", methods=["GET"])
def api_list():
    items = load_items()
    return jsonify({"items": items, "total": format_valor(sum(i["valor"] for i in items))})


@app.route("/api/items", methods=["POST"])
def api_add():
    try:
        peca, nome, valor, url, nota, imagem = validate(request.get_json(silent=True) or {})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    items = load_items()
    items.append({"peca": peca, "nome": nome, "valor": valor, "url": url, "nota": nota, "imagem": imagem})
    save_items(items)
    return jsonify({"ok": True}), 201


@app.route("/api/items/<int:item_id>", methods=["PUT"])
def api_update(item_id):
    items = load_items()
    if item_id < 0 or item_id >= len(items):
        return jsonify({"error": "Item não encontrado."}), 404
    try:
        peca, nome, valor, url, nota, imagem = validate(request.get_json(silent=True) or {})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    antigo = items[item_id]
    registrar_historico(antigo["nome"], antigo["peca"], antigo["valor"], valor)
    items[item_id] = {"peca": peca, "nome": nome, "valor": valor, "url": url, "nota": nota, "imagem": imagem}
    save_items(items)
    return jsonify({"ok": True})


@app.route("/api/items/<int:item_id>", methods=["DELETE"])
def api_delete(item_id):
    items = load_items()
    if item_id < 0 or item_id >= len(items):
        return jsonify({"error": "Item não encontrado."}), 404
    del items[item_id]
    save_items(items)
    return jsonify({"ok": True})


@app.route("/api/historico")
def api_historico():
    return jsonify({"historico": load_historico()})


@app.route("/api/config", methods=["GET"])
def api_get_config():
    return jsonify(load_config())


@app.route("/api/config", methods=["PUT"])
def api_put_config():
    dados = request.get_json(silent=True) or {}
    cfg = load_config()
    cfg["cep"] = re.sub(r"\D", "", str(dados.get("cep", cfg["cep"] or "")))[:8]
    try:
        cfg["orcamento"] = max(0.0, float(str(dados.get("orcamento", cfg["orcamento"] or 0)).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")))
    except (TypeError, ValueError):
        pass
    salvar_config(cfg)
    return jsonify(cfg)


@app.route("/api/ml/status")
def api_ml_status():
    token = ml_token_atual()
    return jsonify({
        "conectado": bool(token),
        "tem_client_id": bool(os.environ.get("ML_CLIENT_ID")),
        "tem_client_secret": bool(os.environ.get("ML_CLIENT_SECRET")),
    })


@app.route("/api/ml/auth", methods=["GET", "POST"])
def api_ml_auth():
    client_id = os.environ.get("ML_CLIENT_ID")
    client_secret = os.environ.get("ML_CLIENT_SECRET")
    redirect_uri = os.environ.get("ML_REDIRECT_URI") or "http://127.0.0.1:5000/api/ml/callback"

    if request.method == "POST":
        code = (request.get_json(silent=True) or {}).get("code") or ""
        if not code.strip():
            return jsonify({"error": "Informe o código de autorização (code)."}), 400
        try:
            resp = http_post_form("https://api.mercadolibre.com/oauth/token", {
                "grant_type": "authorization_code",
                "client_id": client_id,
                "client_secret": client_secret,
                "code": code.strip(),
                "redirect_uri": redirect_uri,
            })
        except Exception:
            return jsonify({"error": "Falha ao trocar o código. Confira client_id/secret e redirect_uri no .env."}), 400
        if not resp.get("access_token"):
            return jsonify({"error": "Resposta inesperada da API.", "detalhe": resp}), 400
        cfg = load_config()
        cfg["ml_token"] = resp["access_token"]
        salvar_config(cfg)
        return jsonify({"ok": True, "user_id": resp.get("user_id")})

    if not client_id or not client_secret:
        return jsonify({"error": "Defina ML_CLIENT_ID e ML_CLIENT_SECRET no arquivo .env (veja .env.example)."}), 400
    auth_url = ("https://auth.mercadolivre.com.br/authorization?"
                f"response_type=code&client_id={client_id}"
                f"&redirect_uri={urllib.parse.quote(redirect_uri, safe='')}")
    return jsonify({"auth_url": auth_url})


@app.route("/api/ml/callback")
def api_ml_callback():
    code = request.args.get("code")
    if not code:
        return "Erro: sem código de autorização.", 400
    try:
        resp = http_post_form("https://api.mercadolibre.com/oauth/token", {
            "grant_type": "authorization_code",
            "client_id": os.environ.get("ML_CLIENT_ID"),
            "client_secret": os.environ.get("ML_CLIENT_SECRET"),
            "code": code,
            "redirect_uri": os.environ.get("ML_REDIRECT_URI") or "http://127.0.0.1:5000/api/ml/callback",
        })
    except Exception:
        return "Falha ao trocar o código. Verifique o .env.", 400
    if not resp.get("access_token"):
        return "Falha na autorização: %s" % resp, 400
    cfg = load_config()
    cfg["ml_token"] = resp["access_token"]
    salvar_config(cfg)
    return "Conectado ao Mercado Livre! Pode fechar esta aba e voltar ao app."


@app.route("/api/precos")
def api_precos():
    q = (request.args.get("q") or "").strip()
    cep = re.sub(r"\D", "", (request.args.get("cep") or ""))[:8] or load_config()["cep"]
    if len(q) < 2:
        return jsonify({"error": "Termo de busca muito curto."}), 400
    resultados, erros = [], []
    fontes = [
        ("Mercado Livre", buscar_mercadolivre_api),
        ("Shopee", buscar_shopee_api),
        ("Busca geral", buscar_ddg),
    ]
    for nome_fonte, funcao in fontes:
        try:
            resultados += funcao(q, cep)
        except Exception:
            erros.append(nome_fonte)
    vistos, unicos = set(), []
    for r in resultados:
        if r["link"] in vistos:
            continue
        vistos.add(r["link"])
        unicos.append(r)
    unicos.sort(key=lambda r: (r["preco"] is None, r["preco"] or 0.0))
    return jsonify({"resultados": unicos, "cep_uso": cep, "erros": erros})


@app.route("/api/export/<formato>")
def api_export(formato):
    items = load_items()
    if formato == "json":
        resp = Response(json.dumps(items, ensure_ascii=False, indent=2), mimetype="application/json; charset=utf-8")
        resp.headers["Content-Disposition"] = "attachment; filename=tabeladecompras.json"
        return resp
    if formato == "csv":
        buf = io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow(HEADERS)
        for item in items:
            writer.writerow([item["peca"], item["nome"], format_valor(item["valor"]), item["url"], item["nota"], item.get("imagem", "")])
        resp = Response("\ufeff" + buf.getvalue(), mimetype="text/csv; charset=utf-8")
        resp.headers["Content-Disposition"] = "attachment; filename=tabeladecompras.csv"
        return resp
    if formato == "xlsx":
        buf = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(HEADERS)
        for item in items:
            ws.append([item["peca"], item["nome"], item["valor"], item["url"], item["nota"], item.get("imagem", "")])
            ws.cell(row=ws.max_row, column=3).number_format = 'R$ #,##0.00'
        wb.save(buf)
        buf.seek(0)
        resp = Response(buf.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp.headers["Content-Disposition"] = "attachment; filename=tabeladecompras.xlsx"
        return resp
    return jsonify({"error": "Formato inválido. Use xlsx, json ou csv."}), 400


@app.route("/api/import", methods=["POST"])
def api_import():
    if "arquivo" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400
    arquivo = request.files["arquivo"]
    nome = arquivo.filename or ""
    if not nome.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx."}), 400
    try:
        items = arquivo_para_items(arquivo.read())
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        return jsonify({"error": "Não foi possível ler a planilha. Verifique se é um .xlsx válido."}), 400
    save_items(items)
    return jsonify({"ok": True, "itens": len(items)})


def find_free_port():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


def port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(("127.0.0.1", port)) == 0


if __name__ == "__main__":
    port = 5000 if not port_in_use(5000) else find_free_port()
    url = f"http://127.0.0.1:{port}"
    print("\n  Meu PC - Lista de Compras")
    print(f"  Abrindo o app em: {url}")
    print("  Pressione Ctrl+C para encerrar.\n")
    if os.environ.get("MEUPC_NO_BROWSER") != "1":
        webbrowser.open(url)
    app.run(debug=False, host="127.0.0.1", port=port)